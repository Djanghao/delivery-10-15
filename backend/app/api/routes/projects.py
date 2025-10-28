from __future__ import annotations

import io
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from ...auth import get_current_user
from ...db import get_db
from ...models import User, ValuableProject
from ...schemas import (
    DeleteByRegionsResponse,
    DeleteProjectsRequest,
    PaginatedProjects,
    ProjectCounts,
    ProjectFull,
    ProjectItem,
)

router = APIRouter(prefix="/api/projects", tags=["projects"])


@router.get("/counts", response_model=ProjectCounts)
def get_project_counts(
    region: str | None = Query(default=None),
    regions: List[str] = Query(default_factory=list),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> ProjectCounts:
    selected = regions or ([region] if region else [])
    base_query = select(func.count()).select_from(ValuableProject)
    if selected:
        base_query = base_query.where(ValuableProject.region_code.in_(selected))

    all_count = int(db.scalar(base_query) or 0)
    parsed_count = int(db.scalar(base_query.where(ValuableProject.parsed_pdf == True)) or 0)
    unparsed_count = int(db.scalar(base_query.where(ValuableProject.parsed_pdf == False).where(ValuableProject.is_invalid == False)) or 0)
    invalid_count = int(db.scalar(base_query.where(ValuableProject.is_invalid == True)) or 0)

    return ProjectCounts(
        all=all_count,
        parsed=parsed_count,
        unparsed=unparsed_count,
        invalid=invalid_count,
    )


@router.get("", response_model=PaginatedProjects)
def list_projects(
    region: str | None = Query(default=None),
    regions: List[str] = Query(default_factory=list),
    parsed: bool | None = Query(default=None),
    invalid: bool | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> PaginatedProjects:
    selected = regions or ([region] if region else [])
    query = select(ValuableProject)
    count_query = select(func.count()).select_from(ValuableProject)
    if selected:
        query = query.where(ValuableProject.region_code.in_(selected))
        count_query = count_query.where(ValuableProject.region_code.in_(selected))
    if invalid is not None:
        query = query.where(ValuableProject.is_invalid == invalid)
        count_query = count_query.where(ValuableProject.is_invalid == invalid)
    elif parsed is not None:
        query = query.where(ValuableProject.parsed_pdf == parsed)
        count_query = count_query.where(ValuableProject.parsed_pdf == parsed)
        if parsed is False:
            query = query.where(ValuableProject.is_invalid == False)
            count_query = count_query.where(ValuableProject.is_invalid == False)
    query = query.order_by(ValuableProject.discovered_at.desc()).offset((page - 1) * size).limit(size)
    items = db.scalars(query).all()
    total = int(db.scalar(count_query) or 0)
    return PaginatedProjects(
        items=[
            ProjectItem(
                projectuuid=item.projectuuid,
                project_name=item.project_name,
                region_code=item.region_code,
                discovered_at=item.discovered_at,
                parsed_pdf=item.parsed_pdf,
                is_invalid=item.is_invalid,
            )
            for item in items
        ],
        total=total,
        page=page,
        size=size,
    )


@router.get("/export")
def export_projects(
    region: str | None = Query(default=None),
    regions: List[str] = Query(default_factory=list),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> Response:
    import json as _json

    selected = regions or ([region] if region else [])
    query = select(ValuableProject).order_by(ValuableProject.discovered_at.desc())
    if selected:
        query = query.where(ValuableProject.region_code.in_(selected))
    projects = db.scalars(query).all()

    parse_field_names = [
        "项目名称",
        "项目类型",
        "建设性质",
        "拟开工时间",
        "拟建成时间",
        "建设规模与建设内容（生产能力）",
        "项目联系人姓名",
        "项目联系人手机",
        "总投资",
        "固定投资",
        "土建工程",
        "设备购置费",
        "安装工程",
        "工程建设其他费用",
        "预备费",
        "建设期利息",
        "铺底流动资金",
        "财政性资金",
        "自有资金（非财政性资金）",
        "银行贷款",
        "其它",
        "项目（法人）单位",
        "成立日期",
        "法定代表人",
        "法定代表人手机号码",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "项目列表"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    basic_headers = ["项目编号", "项目名称", "地区代码", "发现时间", "解析状态"]
    headers = basic_headers + parse_field_names

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    parsed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    unparsed_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, project in enumerate(projects, 2):
        is_parsed = bool(project.parsed_pdf)
        status_fill = parsed_fill if is_parsed else unparsed_fill

        basic_data = [
            project.projectuuid,
            project.project_name,
            project.region_code,
            project.discovered_at.strftime("%Y-%m-%d %H:%M:%S"),
            "已解析" if is_parsed else "未解析",
        ]

        for col_idx, value in enumerate(basic_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment if col_idx in [1, 3, 5] else data_alignment
            if col_idx == 5:
                cell.fill = status_fill
                cell.font = Font(bold=True, color="1F4E78" if is_parsed else "9C6500")

        parsed_data = {}
        if project.pdf_extract_json:
            try:
                parsed_data = _json.loads(project.pdf_extract_json)
            except Exception:
                pass

        for col_offset, field in enumerate(parse_field_names, len(basic_data) + 1):
            value = parsed_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_offset, value=value)
            cell.border = thin_border
            cell.alignment = data_alignment

    column_widths = {
        1: 32,
        2: 40,
        3: 12,
        4: 20,
        5: 12,
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    for col_idx in range(6, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    headers = {"Content-Disposition": "attachment; filename=valuable_projects.xlsx"}
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.delete("", status_code=204)
def delete_projects(payload: DeleteProjectsRequest, db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> Response:
    if not payload.projectuuids:
        raise HTTPException(status_code=400, detail="必须提供要删除的项目uuid 列表")
    stmt = delete(ValuableProject).where(ValuableProject.projectuuid.in_(payload.projectuuids))
    db.execute(stmt)
    db.commit()
    return Response(status_code=204)


@router.delete("/by-regions", response_model=DeleteByRegionsResponse)
def delete_by_regions(regions: List[str] = Query(default_factory=list), db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> DeleteByRegionsResponse:
    if not regions:
        raise HTTPException(status_code=400, detail="必须提供至少一个地区进行删除")
    stmt = delete(ValuableProject).where(ValuableProject.region_code.in_(regions))
    result = db.execute(stmt)
    db.commit()
    deleted = int(result.rowcount or 0)
    return DeleteByRegionsResponse(deleted=deleted)


@router.get("/{projectuuid}", response_model=ProjectFull)
def get_project_full(projectuuid: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)) -> ProjectFull:
    project = db.get(ValuableProject, projectuuid)
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    pdf_extract = None
    if project.pdf_extract_json:
        import json as _json

        try:
            pdf_extract = _json.loads(project.pdf_extract_json)
        except Exception:
            pdf_extract = None
    return ProjectFull(
        projectuuid=project.projectuuid,
        project_name=project.project_name,
        region_code=project.region_code,
        discovered_at=project.discovered_at,
        parsed_pdf=bool(project.parsed_pdf),
        parsed_at=project.parsed_at,
        pdf_extract=pdf_extract,
        pdf_file_path=project.pdf_file_path,
    )
